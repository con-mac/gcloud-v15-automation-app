"""
Questionnaire parser for G-Cloud Capabilities Questionnaire
Parses Excel file and extracts questions grouped by Section Name
"""

import os
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


class QuestionnaireParser:
    """Parses G-Cloud questionnaire Excel file and extracts questions by LOT"""
    
    def __init__(self, excel_path: Optional[str] = None):
        """
        Initialize questionnaire parser
        
        Args:
            excel_path: Path to Excel file (defaults to docs folder)
        """
        if excel_path:
            self.excel_path = Path(excel_path)
        else:
            # Try multiple possible locations
            possible_paths = []
            
            # 1. /app/docs (Azure Functions/Docker standard location)
            possible_paths.append(Path("/app/docs") / "RM1557.15-G-Cloud-question-export (1).xlsx")
            
            # 2. Relative to this file (for local development)
            backend_dir = Path(__file__).parent.parent.parent
            project_root = backend_dir.parent  # Go up one more level to project root
            possible_paths.append(project_root / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx")
            
            # 3. backend/docs
            possible_paths.append(backend_dir / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx")
            
            # 4. Same level as backend (for Azure Functions deployment structure)
            # In Azure Functions, structure might be: /home/site/wwwroot/app/... and /home/site/wwwroot/docs/...
            azure_root = Path("/home/site/wwwroot")
            if azure_root.exists():
                possible_paths.append(azure_root / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx")
            
            # 5. Current working directory/docs
            possible_paths.append(Path.cwd() / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx")
            
            # Find the first existing path
            self.excel_path = None
            for path in possible_paths:
                if path.exists():
                    self.excel_path = path
                    logger.info(f"Found questionnaire Excel file at: {path}")
                    break
            
            if not self.excel_path:
                # Try one more: check if docs folder exists at any level
                for base in [Path("/app"), Path("/home/site/wwwroot"), Path.cwd(), backend_dir, project_root]:
                    test_path = base / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx"
                    if test_path.exists():
                        self.excel_path = test_path
                        logger.info(f"Found questionnaire Excel file at: {test_path}")
                        break
        
        if not self.excel_path.exists():
            # List possible paths for debugging
            possible_paths = [
                project_root / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx",
                backend_dir / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx",
                Path("/app/docs") / "RM1557.15-G-Cloud-question-export (1).xlsx",
                Path(__file__).parent.parent.parent / "docs" / "RM1557.15-G-Cloud-question-export (1).xlsx",
            ]
            logger.error(f"Questionnaire Excel file not found. Tried paths: {[str(p) for p in possible_paths]}")
            raise FileNotFoundError(f"Questionnaire Excel file not found: {self.excel_path}")
    
    def parse_questions_for_lot(self, lot: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        Parse questions for a specific LOT, grouped by Section Name
        
        Args:
            lot: LOT number ("3", "2a", or "2b")
            
        Returns:
            Dict mapping section names to lists of questions
        """
        # Map LOT to sheet name
        lot_sheet_map = {
            '3': 'Services cloud support LOT 3',
            '2a': 'Services Iaas (LOT 2a)',
            '2b': 'Service Saas (LOT 2b)'
        }
        
        sheet_name = lot_sheet_map.get(lot)
        if not sheet_name:
            raise ValueError(f"Invalid LOT: {lot}. Must be '3', '2a', or '2b'")
        
        try:
            # Load with data_only=True to get actual values, but we need to check fills separately
            wb = load_workbook(self.excel_path, data_only=True)
            # Also load without data_only to check fills
            wb_fills = load_workbook(self.excel_path, data_only=False)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
            
            sheet = wb[sheet_name]
            sheet_fills = wb_fills[sheet_name]  # For checking fills
            
            # Get header row
            headers = [cell.value for cell in sheet[1]]
            
            # Find column indices
            section_col_idx = None
            question_col_idx = None
            question_advice_col_idx = None
            question_hint_col_idx = None
            question_type_col_idx = None
            
            for idx, header in enumerate(headers, 1):
                if header:
                    header_lower = str(header).lower()
                    if ('section name' in header_lower or 'section' in header_lower) and section_col_idx is None:
                        section_col_idx = idx
                    elif 'question' in header_lower and 'advice' not in header_lower and 'hint' not in header_lower and 'type' not in header_lower and 'follow up' not in header_lower and question_col_idx is None:
                        question_col_idx = idx
                    elif 'question advice' in header_lower:
                        question_advice_col_idx = idx
                    elif 'question hint' in header_lower:
                        question_hint_col_idx = idx
                    elif 'question type' in header_lower:
                        question_type_col_idx = idx
            
            if not section_col_idx or not question_col_idx:
                raise ValueError("Could not find required columns in Excel file")
            
            # Find answer option columns (Answer 1, Answer 2, etc.)
            answer_cols = []
            for idx, header in enumerate(headers, 1):
                if header and 'answer' in str(header).lower():
                    answer_cols.append(idx)
            
            # Group questions by section
            sections: Dict[str, List[Dict[str, Any]]] = {}
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                # Check if row should be ignored (red fill in column 2)
                # Get the fill cell from the fills sheet
                if len(row) > 1:
                    fill_cell = sheet_fills.cell(row=row_idx, column=2)
                    if fill_cell.fill and isinstance(fill_cell.fill, PatternFill):
                        fill_color = fill_cell.fill.fgColor
                        if fill_color:
                            rgb = fill_color.rgb
                            if rgb:
                                rgb_str = str(rgb).upper()
                                # Check for red fills - common patterns: FFFF0000 (red)
                                # Red typically has high red component
                                if 'FF0000' in rgb_str or (rgb_str.startswith('FF') and len(rgb_str) >= 8):
                                    # Likely red fill, skip this row
                                    continue
                
                # Get section name (row is now values_only, so it's a tuple)
                section_name = row[section_col_idx - 1] if len(row) >= section_col_idx else None
                if not section_name or not str(section_name).strip():
                    continue
                
                section_name = str(section_name).strip()
                # Fix encoding issues
                section_name = section_name.replace('â€™', "'").replace('â€"', '"').replace('â€"', '"')
                
                # Get question text
                question_text = row[question_col_idx - 1] if len(row) >= question_col_idx else None
                if not question_text or not str(question_text).strip():
                    continue
                
                question_text = str(question_text).strip()
                # Fix encoding issues: replace â€™ with '
                question_text = question_text.replace('â€™', "'").replace('â€"', '"').replace('â€"', '"')
                
                # Get question type
                question_type = None
                if question_type_col_idx and len(row) >= question_type_col_idx:
                    qtype_value = row[question_type_col_idx - 1]
                    if qtype_value:
                        question_type = str(qtype_value).strip()
                
                # Get question advice and hint
                question_advice = None
                if question_advice_col_idx and len(row) >= question_advice_col_idx:
                    advice_value = row[question_advice_col_idx - 1]
                    if advice_value:
                        question_advice = str(advice_value).strip()
                        # Fix encoding issues
                        question_advice = question_advice.replace('â€™', "'").replace('â€"', '"').replace('â€"', '"')
                
                question_hint = None
                if question_hint_col_idx and len(row) >= question_hint_col_idx:
                    hint_value = row[question_hint_col_idx - 1]
                    if hint_value:
                        question_hint = str(hint_value).strip()
                        # Fix encoding issues
                        question_hint = question_hint.replace('â€™', "'").replace('â€"', '"').replace('â€"', '"')
                
                # Get answer options
                answer_options = []
                for col_idx in answer_cols:
                    if len(row) >= col_idx:
                        answer_value = row[col_idx - 1]
                    if answer_value and str(answer_value).strip():
                        answer_option = str(answer_value).strip()
                        # Fix encoding issues
                        answer_option = answer_option.replace('â€™', "'").replace('â€"', '"').replace('â€"', '"')
                        answer_options.append(answer_option)
                
                # Normalize question type
                normalized_type = self._normalize_question_type(question_type, answer_options)
                
                # Create question object
                question = {
                    'question_text': question_text,
                    'question_type': normalized_type,
                    'question_advice': question_advice,
                    'question_hint': question_hint,
                    'answer_options': answer_options if answer_options else None,
                    'row_index': row_idx
                }
                
                # Add to section
                if section_name not in sections:
                    sections[section_name] = []
                sections[section_name].append(question)
            
            return sections
            
        except Exception as e:
            logger.error(f"Error parsing questionnaire for LOT {lot}: {e}", exc_info=True)
            raise
    
    def _normalize_question_type(self, question_type: Optional[str], answer_options: List[str]) -> str:
        """
        Normalize question type to standard values
        
        Args:
            question_type: Raw question type from Excel
            answer_options: List of answer options
            
        Returns:
            Normalized question type: 'radio', 'checkbox', 'text', 'textarea', 'list'
        """
        if not question_type:
            # Infer from answer options
            if answer_options:
                return 'radio'  # Default to radio if options exist
            return 'text'
        
        qtype_lower = question_type.lower()
        
        if 'radio' in qtype_lower or 'single' in qtype_lower:
            return 'radio'
        elif 'checkbox' in qtype_lower or 'multiple' in qtype_lower or 'grouped checkbox' in qtype_lower:
            return 'checkbox'
        elif 'textarea' in qtype_lower or 'text area' in qtype_lower:
            return 'textarea'
        elif 'list' in qtype_lower or 'list of text' in qtype_lower:
            return 'list'
        elif 'text field' in qtype_lower or 'textfield' in qtype_lower:
            return 'text'
        else:
            # Default based on answer options
            if answer_options:
                return 'radio'
            return 'text'
    
    def get_sections_for_lot(self, lot: str) -> List[str]:
        """
        Get list of section names for a LOT (in order)
        
        Args:
            lot: LOT number ("3", "2a", or "2b")
            
        Returns:
            List of section names in order they appear
        """
        questions = self.parse_questions_for_lot(lot)
        # Return sections in order they first appear
        section_order = []
        seen = set()
        
        # We need to preserve order, so we'll parse again and track order
        lot_sheet_map = {
            '3': 'Services cloud support LOT 3',
            '2a': 'Services Iaas (LOT 2a)',
            '2b': 'Service Saas (LOT 2b)'
        }
        
        sheet_name = lot_sheet_map.get(lot)
        if not sheet_name:
            return []
        
        try:
            wb = load_workbook(self.excel_path, data_only=True)
            sheet = wb[sheet_name]
            
            headers = [cell.value for cell in sheet[1]]
            section_col_idx = None
            for idx, header in enumerate(headers, 1):
                if header and ('section name' in str(header).lower() or 'section' in str(header).lower()):
                    section_col_idx = idx
                    break
            
            if not section_col_idx:
                return list(questions.keys())
            
            for row in sheet.iter_rows(min_row=2, values_only=False):
                if len(row) > 1:
                    cell = row[1]
                    if cell.fill and isinstance(cell.fill, PatternFill):
                        fill_color = cell.fill.fgColor
                        if fill_color and fill_color.rgb and 'FF' in str(fill_color.rgb).upper():
                            continue
                
                if len(row) >= section_col_idx:
                    section_name = row[section_col_idx - 1].value
                    if section_name and str(section_name).strip():
                        section_name = str(section_name).strip()
                        if section_name not in seen:
                            section_order.append(section_name)
                            seen.add(section_name)
            
            return section_order
            
        except Exception as e:
            logger.error(f"Error getting sections for LOT {lot}: {e}")
            return list(questions.keys())

